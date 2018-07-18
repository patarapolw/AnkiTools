import openpyxl
from random import choice
import logging
import importlib_resources

debugger_logger = logging.getLogger('debug')


class ExcelExport:
    def __init__(self,
                 out_file: str,
                 raw_excel_data: dict=None,
                 excel_formatting_file: str=None):
        try:
            self.wb = openpyxl.load_workbook(excel_formatting_file)
        except (FileNotFoundError, TypeError):
            with importlib_resources.path('AnkiTools.excel', 'template.xlsx') as f:
                self.wb = openpyxl.load_workbook(f)

        sheet_names = self.wb.sheetnames
        debugger_logger.debug(sheet_names)

        if '.settings' not in sheet_names:
            self.wb.create_sheet('.settings', 0)
        else:
            sheet_names.remove('.settings')

        for row in self.wb['.settings']:
            for cell in row:
                cell.value = None

        debugger_logger.debug(sheet_names)

        random_sheet = ''
        if '.template' not in sheet_names:
            if len(sheet_names) > 0:
                random_sheet = choice(sheet_names)
                self.wb[choice(sheet_names)].title = '.template'
                debugger_logger.debug('Renamed ' + random_sheet)
                sheet_names.remove(random_sheet)
            else:
                self.wb.create_sheet('.template', 1)
        else:
            sheet_names.remove('.template')

        debugger_logger.debug(sheet_names)

        for sheet_name in sheet_names:
            if sheet_name != random_sheet:
                debugger_logger.debug('Removing ' + sheet_name)
                # self.wb.remove(self.wb[sheet_name])

        if raw_excel_data:
            self.populate_data(raw_excel_data)

        self.out_file = out_file

    def save(self):
        self.wb.active = 0
        self.wb.active.sheet_view.selection[0].activeCell = 'A1'

        self.wb.save(self.out_file)

    def populate_data(self, raw_excel_data: dict):
        formatted_sheet_names = self.wb.sheetnames

        for sheet_name, cell_matrix in raw_excel_data.items():
            if sheet_name not in self.wb.sheetnames:
                self.wb.create_sheet(sheet_name)
            ws = self.wb[sheet_name]
            for row_num, row in enumerate(cell_matrix):
                for col_num, value in enumerate(row):
                    if not isinstance(value, (int, str)):
                        print(value)

                    ws.cell(column=(col_num + 1),
                            row=(row_num + 1),
                            value=value)

        for sheet_name in raw_excel_data.keys():
            if sheet_name not in formatted_sheet_names:
                if '.template' in self.wb.sheetnames:
                    ws = self.wb[sheet_name]
                    self.copy_styles(src=self.wb['.template'], dst=ws)

        if '.template' in self.wb.sheetnames:
            self.wb.remove(self.wb['.template'])

        for sheet_name in self.wb.sheetnames:
            iter_rows = self.wb[sheet_name].iter_rows()
            try:
                next(iter_rows)
                if all([not cell.value for cell in next(iter_rows)]):
                    print(raw_excel_data[sheet_name])
                    self.wb.remove(self.wb[sheet_name])
                    break
            except StopIteration:
                self.wb.remove(self.wb[sheet_name])
                break

    @staticmethod
    def copy_styles(src, dst):
        """

        :param openpyxl.worksheet.worksheet.Worksheet src:
        :param openpyxl.worksheet.worksheet.Worksheet dst:
        :return:

        TODO: load default formattings from a *.xlsx file
        TODO: verify that the styles are correctly formatted
        """
        params = dict()
        for col_letter, col_data in src.column_dimensions.items():
            params[col_letter] = col_data.__dict__

        for col_letter, col_format in params.items():
            dst.column_dimensions[col_letter].__dict__ = col_format
