import openpyxl

from AnkiTools.dir import excel_path


class ExcelFormatter:
    def __init__(self,
                 excel_formatting_file: str=None,
                 excel_data_file: str=None,
                 out_file: str=None):
        try:
            self.wb_formatting = openpyxl.load_workbook(excel_formatting_file)
        except (FileNotFoundError, TypeError):
            self.wb_formatting = openpyxl.load_workbook(excel_path('default.xlsx'))

        try:
            self.wb_data = openpyxl.load_workbook(excel_data_file)
        except (FileNotFoundError, TypeError):
            self.wb_data = None

        if out_file is not None:
            self.out_file = out_file
        else:
            self.out_file = excel_data_file

    def save(self):
        self.wb_data.save(self.out_file)

    def do_formatting(self):
        formatting_sheet_names = self.wb_formatting.sheetnames
        data_sheet_names = self.wb_data.sheetnames
        for sheet_name in data_sheet_names:
            if sheet_name in formatting_sheet_names:
                self.copy_styles(src=self.wb_data[sheet_name],
                                 dst=self.wb_formatting[sheet_name])
            else:
                if '.default' in formatting_sheet_names:
                    self.copy_styles(src=self.wb_data[sheet_name],
                                     dst=self.wb_formatting['.default'])

    @staticmethod
    def copy_styles(src, dst):
        """

        :param openpyxl.worksheet.worksheet.Worksheet src:
        :param openpyxl.worksheet.worksheet.Worksheet dst:
        :return:

        TODO: load default formattings from a *.xlsx file
        TODO: verify that the styles are correctly formatted
        """
        print('Hello, too.')

        dst.column_dimensions['A'].width = src.column_dimensions['A'].width
        dst.column_dimensions['B'].width = src.column_dimensions['B'].width
