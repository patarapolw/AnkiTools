from AnkiTools.tools.edit import editAnki2
from AnkiTools.tools.read import readApkg, readAnki2

import sqlite3, re
import csv
from openpyxl import Workbook
from zipfile import ZipFile
import os


def apkg2anki(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.anki2'

    with ZipFile(fin) as zp:
        with open(fout, 'wb') as f:
            f.write(zp.read('collection.anki2'))


def anki2apkg(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.apkg'

    with editAnki2(fin) as anki:
        anki.export(fout)


def apkg2xlsx(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.xlsx'

    with readApkg(fin) as anki:
        _2xlsx(anki, fout)


def anki2xlsx(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.xlsx'

    with readAnki2(fin) as anki:
        _2xlsx(anki, fout)


def _2xlsx(anki, fout):
    workbook = Workbook()

    sheets = []
    for note in anki.notes.values():
        model = note['model']
        if model['name'] not in sheets:
            worksheet = workbook.create_sheet(model['name'])
            sheets.append(model['name'])

            for i, field in enumerate(model['fields']):
                worksheet.cell(row=1, column=i + 1, value=field)
        else:
            worksheet = workbook[model['name']]

        max_row = worksheet.max_row
        for i, data in enumerate(note['content']):
            worksheet.cell(row=max_row + 1, column=i + 1, value=data)

    workbook.save(fout)


def apkg2tsv(fin, fout_dir=''):
    if fout_dir == '':
        fout_dir = os.path.splitext(fin)[0]
    if not os.path.exists(fout_dir):
        os.mkdir(fout_dir)

    with readApkg(fin) as anki:
        _2tsv(anki, fout_dir)


def anki2tsv(fin, fout_dir=''):
    if fout_dir == '':
        fout_dir = os.path.splitext(fin)[0]
    if not os.path.exists(fout_dir):
        os.mkdir(fout_dir)

    with readAnki2(fin) as anki:
        _2tsv(anki, fout_dir)


def _2tsv(anki, fout_dir):
    fout_tsv = dict()
    for note in anki.notes.values():
        model = note['model']
        if model['name'] not in fout_tsv.keys():
            fout_tsv[model['name']] = open(os.path.join(fout_dir, model['name'] + '.tsv'), 'w')

            for field in model['fields']:
                fout_tsv[model['name']].write(field + '\t')
            fout_tsv[model['name']].write('\n')

        for data in note['content']:
            fout_tsv[model['name']].write(data + '\t')
        fout_tsv[model['name']].write('\n')

    for fout in fout_tsv.values():
        fout.close()


def apkg2csv(fin, fout_dir=''):
    if fout_dir == '':
        fout_dir = os.path.splitext(fin)[0]
    if not os.path.exists(fout_dir):
        os.mkdir(fout_dir)

    with readApkg(fin) as anki:
        _2csv(anki, fout_dir)


def anki2csv(fin, fout_dir=''):
    if fout_dir == '':
        fout_dir = os.path.splitext(fin)[0]
    if not os.path.exists(fout_dir):
        os.mkdir(fout_dir)

    with readAnki2(fin) as anki:
        _2csv(anki, fout_dir)


def _2csv(anki, fout_dir):
    fout_csv = dict()
    writer_csv = dict()
    for note in anki.notes.values():
        model = note['model']
        if model['name'] not in fout_csv.keys():
            fout_csv[model['name']] = open(os.path.join(fout_dir, model['name'] + '.csv'), 'w')
            writer_csv[model['name']] = csv.writer(fout_csv[model['name']])

            writer_csv[model['name']].writerow(model['fields'])

        writer_csv[model['name']].writerow(note['content'])

    for fout in fout_csv.values():
        fout.close()


def apkg2sqlite(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.sqlite'

    with readApkg(fin) as anki:
        _2sqlite(anki, fout)


def anki2sqlite(fin, fout=''):
    if fout == '':
        fout = os.path.splitext(fin)[0] + '.sqlite'

    with readAnki2(fin) as anki:
        _2sqlite(anki, fout)


def _2sqlite(anki, fout):
    with sqlite3.connect(fout) as conn:
        used_models = []
        for note in anki.notes.values():
            model = note['model']
            if model['name'] not in used_models:
                used_models.append(model['name'])

                query = 'CREATE TABLE IF NOT EXISTS {} '.format(_slugify(model['name']))
                query += '(id INTEGER PRIMARY KEY AUTOINCREMENT, '
                for field in model['fields']:
                    query += '{} TEXT, '.format(_slugify(field))
                query =  query[:-2] + ');'

                conn.execute(query)

            query1 = 'INSERT INTO {} ('.format(_slugify(model['name']))
            query2 = ') VALUES ('
            for field in model['fields']:
                query1 += '{}, '.format(_slugify(field))
                query2 += '?, '
            query = query1[:-2] + query2[:-2] + ');'

            conn.execute(query, tuple(note['content']))
            conn.commit()


def _slugify(text, lower=1):
    if lower == 1:
        text = text.strip().lower()
    text = re.sub(r'[^\w _-]+', '', text)
    text = re.sub(r'[- ]+', '_', text)
    return text
