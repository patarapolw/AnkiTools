import openpyxl as px
from collections import OrderedDict, namedtuple
from datetime import datetime

from AnkiTools.AnkiDirect import AnkiDirect
from AnkiTools.tools.defaults import DEFAULT_API_MODEL_DEFINITION

DeckTuple = namedtuple('DeckTuple', ['deck_id', 'deck_name'])
CardTuple = namedtuple('CardTuple', ['card_id', 'note_id', 'deck_name', 'template_order'])


class AnkiExcelSync:
    SHEET_SETTINGS = '.settings'
    SHEET_DECKS = '.decks'

    def __init__(self, excel: str, anki_database: str):
        self.anki_direct = AnkiDirect(anki_database=anki_database)
        self.excel_filename = excel

        self.settings = {
            'models': dict(),
            'decks': dict()
        }

        try:
            self.wb = px.load_workbook(self.excel_filename)
        except FileNotFoundError:
            self.wb = self.create()

    def to_excel(self):
        self.wb.save(self.excel_filename)

    def to_sqlite(self):
        self.anki_direct.add(self.to_json())

    def to_json(self):
        payload = {
            'data': dict(),
            'definitions': dict()
        }

        sheet_names = self.wb.sheetnames
        try:
            sheet_names.remove(self.SHEET_SETTINGS)
            sheet_names.remove(self.SHEET_DECKS)
        except ValueError:
            pass

        for sheet_name in sheet_names:
            payload['data'][sheet_name] = list()

            row_iter = self.wb[sheet_name].iter_rows()
            header = [cell.value for cell in next(row_iter)]
            for row in row_iter:
                record = OrderedDict(zip(header, [cell.value for cell in row]))
                formatted_record = {
                    'data': record,
                    'decks': {
                        'Card 1': sheet_name
                    }
                }

                payload['data'][sheet_name].append(formatted_record)

            payload['definitions'][sheet_name] = DEFAULT_API_MODEL_DEFINITION
            payload['definitions'][sheet_name]['templates'][0]['data']['qfmt'] = \
                payload['definitions'][sheet_name]['templates'][0]['data']['qfmt'] % header[0]
            payload['definitions'][sheet_name]['templates'][0]['data']['afmt'] = \
                payload['definitions'][sheet_name]['templates'][0]['data']['afmt'] % header[1]

        return payload

    def create(self):
        wb = px.Workbook()

        ws = wb.active
        ws.title = self.SHEET_SETTINGS
        ws.append(['Created', datetime.fromtimestamp(datetime.now().timestamp()).isoformat()])
        ws.append(['Modified', datetime.fromtimestamp(datetime.now().timestamp()).isoformat()])

        # Getting sheet names
        models = self.anki_direct.models_dict
        model_id_to_name = dict()
        for model_id, model_dict in models.items():
            sheet_name = model_dict['name']
            print('Creating sheet {}'.format(sheet_name))

            # Writing header
            if sheet_name not in wb.sheetnames:
                header = ['id']
                field_pairs = [(fld['ord'], fld['name']) for fld in model_dict['flds']]
                header.extend([x[1] for x in sorted(field_pairs)])
                header.append('Tags')

                wb.create_sheet(sheet_name)
                wb[sheet_name].append(header)

            model_id_to_name[model_id] = sheet_name
            self.settings['models'][sheet_name] = {
                'id': model_id,
                'templates': model_dict['tmpls']
            }

        # Getting sheet contents
        notes_iter = self.anki_direct.notes
        for note in notes_iter:
            try:
                sheet_name = model_id_to_name[str(note['mid'])]
            except KeyError:
                continue

            # Writing record
            print('Creating note {}'.format(note['id']))

            record = [note['id']]
            record.extend(note['formatted_flds'])
            record.append(note['tags'])
            wb[sheet_name].append(record)

        # Getting deck id and names
        decks_dict = self.anki_direct.decks_dict
        for deck_info in decks_dict.values():
            self.settings['decks'][deck_info['name']] = deck_info

        # Getting card distribution
        wb.create_sheet(self.SHEET_DECKS, 1)

        wb[self.SHEET_DECKS].append(CardTuple._fields)
        cards_iter = self.anki_direct.cards
        for card in cards_iter:
            record = CardTuple(
                card_id=card['id'],
                note_id=card['nid'],
                deck_name=decks_dict[str(card['did'])]['name'],
                template_order=card['ord']
            )
            wb[self.SHEET_DECKS].append(record)

        return wb
