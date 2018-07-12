import openpyxl as px
from openpyxl.utils import get_column_letter

from collections import OrderedDict, namedtuple
from time import time
from datetime import datetime
import json
import logging

from .api.ankidirect import AnkiDirect
from .tools.defaults import DEFAULT_API_MODEL_DEFINITION

DeckTuple = namedtuple('DeckTuple', ['deck_id', 'deck_name'])
CardTuple = namedtuple('CardTuple', ['card_id', 'note_id', 'deck_name', 'template_order'])

COLUMN_ID_MIN_WIDTH = len(str(int(time() * 1000))) + 3
COLUMN_TIMESTAMP_WIDTH = len(datetime.fromtimestamp(datetime.now().timestamp()).isoformat()) + 1


class AnkiExcelSync:
    SHEET_SETTINGS = '.settings'
    SHEET_DECKS = '.decks'

    def __init__(self, excel: str, anki_database: str):
        self.anki_direct = AnkiDirect(anki_database=anki_database)
        self.excel_filename = excel

        self.settings = {
            'models': dict(),
            'decks': dict()
        }

        try:
            self.wb = px.load_workbook(self.excel_filename)
        except FileNotFoundError:
            self.wb = self.create()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        self.close()

    def close(self):
        self.wb.close()

    def save(self):
        self.wb.save(self.excel_filename)

    def to_excel(self):
        self.wb.save(self.excel_filename)

    def to_sqlite(self):
        self.anki_direct.add(self.to_json())

    def to_json(self):
        payload = {
            'data': dict(),
            'definitions': dict()
        }

        sheet_names = self.wb.sheetnames
        try:
            sheet_names.remove(self.SHEET_SETTINGS)
            sheet_names.remove(self.SHEET_DECKS)
        except ValueError:
            pass

        for sheet_name in sheet_names:
            payload['data'][sheet_name] = list()

            row_iter = self.wb[sheet_name].iter_rows()
            header = list(self.get_cell_value_iter(next(row_iter)))
            for row in row_iter:
                record = OrderedDict(zip(header, self.get_cell_value_iter(row)))
                formatted_record = {
                    'data': record,
                    'decks': {
                        'Card 1': sheet_name
                    }
                }

                payload['data'][sheet_name].append(formatted_record)

            # This will further be "string-formatted", so it needs to be deep-copied.
            # Currently implemented using a ReadOnlyJsonObject object.
            payload['definitions'][sheet_name] = DEFAULT_API_MODEL_DEFINITION.to_json_object()

            payload['definitions'][sheet_name]['templates'][0]['data']['qfmt'] = \
                payload['definitions'][sheet_name]['templates'][0]['data']['qfmt'] % header[0]
            payload['definitions'][sheet_name]['templates'][0]['data']['afmt'] = \
                payload['definitions'][sheet_name]['templates'][0]['data']['afmt'] % header[1]

        return payload

    def create(self):
        wb = px.Workbook()

        ws = wb.active
        ws.title = self.SHEET_SETTINGS
        ws.column_dimensions['B'].width = COLUMN_TIMESTAMP_WIDTH
        timestamp = datetime.fromtimestamp(datetime.now().timestamp()).isoformat()
        ws.append(['Created', timestamp])
        ws.append(['Modified', timestamp])

        # Getting sheet names
        models = self.anki_direct.models_dict
        model_id_to_name = dict()
        for model_id, model_dict in models.items():
            sheet_name = model_dict['name']
            logging.info('Creating sheet {}'.format(sheet_name))

            # Writing header
            if sheet_name not in wb.sheetnames:
                header = ['id']
                field_pairs = [(fld['ord'], fld['name']) for fld in model_dict['flds']]
                header.extend([x[1] for x in sorted(field_pairs)])
                header.append('Tags')

                wb.create_sheet(sheet_name)
                ws = wb[sheet_name]
                ws.append(header)

                ws.column_dimensions['A'].width = COLUMN_ID_MIN_WIDTH
                for header_id in range(1, len(header)):
                    width = len(header[header_id]) * 1.2
                    if width < 15:
                        width = 15
                    ws.column_dimensions[get_column_letter(header_id + 1)].width = width

            model_id_to_name[model_id] = sheet_name
            self.settings['models'][sheet_name] = {
                'id': model_id,
                'templates': model_dict['tmpls']
            }

        # Getting sheet contents
        notes_iter = self.anki_direct.notes
        for note in notes_iter:
            try:
                sheet_name = model_id_to_name[str(note['mid'])]
            except KeyError:
                continue

            # Writing record
            logging.info('Creating note {} - {}'.format(note['id'],
                                                        json.dumps(note['formatted_flds'], ensure_ascii=False)))

            record = [note['id']]
            record.extend(note['formatted_flds'])
            record.append(note['tags'])
            wb[sheet_name].append(record)

        # Getting deck id and names
        decks_dict = self.anki_direct.decks_dict
        for deck_info in decks_dict.values():
            self.settings['decks'][deck_info['name']] = deck_info

        # Getting card distribution
        wb.create_sheet(self.SHEET_DECKS, 1)
        ws = wb[self.SHEET_DECKS]
        ws.append(CardTuple._fields)
        ws.column_dimensions['A'].width = COLUMN_ID_MIN_WIDTH
        ws.column_dimensions['B'].width = COLUMN_ID_MIN_WIDTH
        for i in range(2, len(CardTuple._fields) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        cards_iter = self.anki_direct.cards
        for card in cards_iter:
            record = CardTuple(
                card_id=card['id'],
                note_id=card['nid'],
                deck_name=decks_dict[str(card['did'])]['name'],
                template_order=card['ord']
            )
            ws.append(record)

        # Delete empty sheets
        sheet_names = wb.sheetnames
        sheet_names.remove(self.SHEET_SETTINGS)
        sheet_names.remove(self.SHEET_DECKS)

        for sheet_name in sheet_names:
            if wb[sheet_name].max_row <= 1:
                wb.remove(wb[sheet_name])

        return wb

    @staticmethod
    def get_cell_value_iter(cell_iter):
        for cell in cell_iter:
            value = cell.value
            if not value:
                yield ''
            else:
                yield value
